from __future__ import annotations

from io import BytesIO

import openpyxl

from app.domain.entities import EstimateRowStatus, NodeMatch


def _make_original(rows: list[tuple[int, str, str]]) -> bytes:
    """rows: (physical_row, code-в-колонке-A, имя). Строка 1 — заголовки."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="№")
    ws.cell(row=1, column=2, value="Наименование")
    ws.cell(row=1, column=3, value="Статья СМР")  # пустая колонка-приёмник
    for phys, code, name in rows:
        ws.cell(row=phys, column=1, value=code)
        ws.cell(row=phys, column=2, value=name)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_export_writes_final_code_to_node_physrow(
    client, auth_headers, estimate_repo, object_storage, seed_estimate_with_source_index
):
    # узел с source_index=33 → физ.строка 35
    eid, nid = seed_estimate_with_source_index(source_index=33)
    object_storage.store["key"] = _make_original([(35, "1.1.5", "Кладка")])
    estimate_repo.nodes[nid]["embedding"] = [0.1]
    estimate_repo.save_node_match(
        nid, NodeMatch(EstimateRowStatus.NEEDS_REVIEW, matched_id=7, matched_code="2.1",
                       matched_name="Статья", score=0.7, candidates=[]),
    )
    estimate_repo.save_review_decision(
        nid, review_status="overridden", final_article_id=7,
        final_code="ИТ-9", final_name="Выбрано",
    )
    resp = client.get(f"/api/estimates/{eid}/export", headers=auth_headers)
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    # колонка «Статья СМР»: формат «(код) Название»
    assert ws.cell(row=35, column=3).value == "(ИТ-9) Выбрано"


def test_export_unreviewed_needs_review_is_blank(
    client, auth_headers, estimate_repo, object_storage, seed_estimate_with_source_index
):
    eid, nid = seed_estimate_with_source_index(source_index=0)
    object_storage.store["key"] = _make_original([(2, "1", "Узел")])
    estimate_repo.nodes[nid]["embedding"] = [0.1]
    estimate_repo.save_node_match(
        nid, NodeMatch(EstimateRowStatus.NEEDS_REVIEW, matched_id=7, matched_code="2.1",
                       matched_name="Статья", score=0.7, candidates=[]),
    )
    resp = client.get(f"/api/estimates/{eid}/export", headers=auth_headers)
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    assert wb.active.cell(row=2, column=3).value in (None, "")  # пусто, не AI-догадка


def test_export_strict_409_when_unreviewed(
    client, auth_headers, estimate_repo, object_storage, seed_estimate_with_source_index
):
    eid, nid = seed_estimate_with_source_index(source_index=0)
    object_storage.store["key"] = _make_original([(2, "1", "Узел")])
    estimate_repo.nodes[nid]["embedding"] = [0.1]
    estimate_repo.save_node_match(
        nid, NodeMatch(EstimateRowStatus.NO_MATCH, candidates=[]),
    )
    resp = client.get(f"/api/estimates/{eid}/export?strict=true", headers=auth_headers)
    assert resp.status_code == 409


def test_export_storage_down_503(
    client, auth_headers, estimate_repo, object_storage, seed_estimate_with_source_index
):
    eid, nid = seed_estimate_with_source_index(source_index=0)
    # ключа нет в store → FakeObjectStorage.get кидает StorageError
    # (контракт порта ObjectStorage.get, как реальный S3-адаптер) → роут отвечает 503, не 500.
    resp = client.get(f"/api/estimates/{eid}/export", headers=auth_headers)
    assert resp.status_code == 503
