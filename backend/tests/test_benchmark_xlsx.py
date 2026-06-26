from __future__ import annotations

import openpyxl

from app.domain.benchmark import BenchmarkKind
from app.infrastructure.benchmark_xlsx import read_benchmark_nodes
from app.services.estimate_parser import EstimateParser


def _make_xlsx(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["№ раздела", "Статья СМР", "Наименование раздела / позиции"])
    for r in rows:
        ws.append(r)
    wb.save(path)


def test_read_nodes_assigns_kinds(tmp_path):
    p = tmp_path / "gold.xlsx"
    _make_xlsx(p, [
        ["1", "(1) Подготовительные работы", "Подготовительные работы"],  # matchable
        ["1.1", None, "1 Этап ЖК"],                                       # structural
        ["10", None, "Инженерные системы"],                               # no_article
        [None, None, "листовая позиция"],                                 # пропуск (не узел)
    ])
    nodes = read_benchmark_nodes(str(p))
    by_code = {n.code: n for n in nodes}
    assert set(by_code) == {"1", "1.1", "10"}
    assert by_code["1"].expected_kind == BenchmarkKind.MATCHABLE.value
    assert by_code["1"].expected_article_code == "1"
    assert by_code["1.1"].expected_kind == BenchmarkKind.STRUCTURAL.value
    assert by_code["10"].expected_kind == BenchmarkKind.NO_ARTICLE.value
    assert nodes[0].source_index < nodes[1].source_index  # порядок сохранён


def test_name_cleaning_matches_parser(tmp_path):
    # Паритет имён: benchmark_xlsx._clean и EstimateParser._clean_name — две копии,
    # могут разойтись. Имя узла кормит крошку в проде, поэтому чистка ОБЯЗАНА совпадать,
    # иначе бенчмарк хранит не тот текст, что эмбеддится в проде (дыра в достоверности).
    p = tmp_path / "gold.xlsx"
    _make_xlsx(p, [
        ["1", "(1) Подготовительные", "Подготовительные  работы\xa0и содержание"],
        ["1.1", None, "Мобилизация  площадки"],
    ])
    seeds = read_benchmark_nodes(str(p))
    parsed = EstimateParser().parse(p.read_bytes())
    assert {(s.code, s.name) for s in seeds} == {(n.code, n.name) for n in parsed.nodes}


def test_coded_row_with_empty_name_skipped_like_parser(tmp_path):
    # Строка с валидным кодом, но пустой ячейкой имени: openpyxl отдаёт None.
    # Без None→"" в _clean она осела бы как name="None"; EstimateParser (NaN→"nan")
    # такую строку отбрасывает. Проверяем, что оба читателя дают одинаковый набор.
    p = tmp_path / "gold.xlsx"
    _make_xlsx(p, [
        ["1", "(1) Подготовительные", "Подготовительные работы"],
        ["2", None, None],  # код есть, имя пустое → должна отсеяться
    ])
    seeds = read_benchmark_nodes(str(p))
    assert {s.code for s in seeds} == {"1"}
    parsed = EstimateParser().parse(p.read_bytes())
    assert {(s.code, s.name) for s in seeds} == {(n.code, n.name) for n in parsed.nodes}


def test_numeric_typed_codes_match_parser(tmp_path):
    p = tmp_path / "numeric.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["№ раздела", "Статья СМР", "Наименование раздела / позиции"])
    ws.append([1, "(1) Подготовительные", "Подготовительные работы"])     # int
    ws.append([10, None, "Инженерные системы"])                           # int
    ws.cell(row=4, column=1, value=1.1)                                   # float
    ws.cell(row=4, column=3, value="Мобилизация")
    wb.save(p)
    seeds = read_benchmark_nodes(str(p))
    parsed = EstimateParser().parse(p.read_bytes())
    assert {(s.code, s.name) for s in seeds} == {(n.code, n.name) for n in parsed.nodes}
