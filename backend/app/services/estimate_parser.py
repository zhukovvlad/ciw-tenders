"""Иерархический парсер сметы. Чистая логика: bytes → ParsedEstimate. Без БД/AI."""

from __future__ import annotations

import io
import re

import pandas as pd
from openpyxl import load_workbook

from app.domain.classification import detect_structural_anomalies, resolve_ancestor_indices
from app.domain.entities import EstimateNode, EstimatePosition, ParsedEstimate

SECTION_NO_COLUMN = "№ раздела"
NAME_COLUMN = "Наименование раздела / позиции"
SECTION_TYPE_COLUMN = "Вид раздела"


def _clean_name(value: object) -> str:
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def _norm_code(value: object) -> str | None:
    """Нормализует код из openpyxl-ячейки к строковому виду для сверки с pandas.

    openpyxl читает числовые ячейки как int/float: 1 → 1 или 1.0 → нужно «1»,
    а не «1.0» (иначе сверка с pandas-строкой «1» всегда расходилась бы).
    """
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():  # openpyxl numeric cell 1.0 → "1"
        value = int(value)
    code = re.sub(r"\s+", "", str(value)).strip(".")
    return code or None


class EstimateParser:
    """Строит дерево узлов из «№ раздела»; листья (№=NaN) — контекст."""

    def parse(self, content: bytes) -> ParsedEstimate:
        # № раздела — принудительно строкой: иначе number-formatted ячейка
        # коэрсится во float (1 → 1.0 → два сегмента), 1.10 схлопывается в 1.1.
        df = pd.read_excel(io.BytesIO(content), engine="openpyxl", dtype={SECTION_NO_COLUMN: str})
        missing = {SECTION_NO_COLUMN, NAME_COLUMN} - set(df.columns)
        if missing:
            raise ValueError(f"В файле отсутствуют обязательные колонки: {sorted(missing)}")

        outline_by_si, code_by_si = self._read_outline(content)

        warnings: list[str] = []
        positions: list[EstimatePosition] = []
        last_node_code: str | None = None
        outline_desync = False  # рассинхрон pandas↔openpyxl → outline-детекция недоступна
        # coded-узлы (только они идут в стек глубины) — собираем в порядке документа
        coded: list[dict] = []  # {source_index, code, name, segments, section_type, outline}
        top_type_by_segment: dict[str, str | None] = {}

        # source_index = ИСХОДНАЯ 0-based позиция (df.iterrows сохраняет RangeIndex);
        # НЕ enumerate по выжившим, НЕ reset_index — иначе после skip уедет на -1.
        for raw_idx, record in df.iterrows():
            si = int(raw_idx)  # type: ignore[arg-type]
            no = record[SECTION_NO_COLUMN]
            name = _clean_name(record[NAME_COLUMN])
            if not name or name.lower() == "nan":
                warnings.append(f"строка {si}: пустое имя — пропущена")
                continue

            if pd.isna(no):  # POSITION
                if last_node_code is None:
                    warnings.append(f"строка {si}: позиция до первого узла")
                positions.append(EstimatePosition(name, last_node_code, si))
                continue

            code = re.sub(r"\s+", "", str(no)).strip(".")
            segments = code.split(".")
            if not code or not all(seg.isdigit() for seg in segments):
                warnings.append(f"строка {si}: нечисловой код '{no}' → позиция")
                positions.append(EstimatePosition(name, last_node_code, si))
                continue

            # Сверка выравнивания pandas↔openpyxl (v2: warning, НЕ raise — outline только детекция)
            if _norm_code(code_by_si.get(si)) != code:
                outline_desync = True

            if len(segments) == 1:
                vid = record[SECTION_TYPE_COLUMN] if SECTION_TYPE_COLUMN in df.columns else None
                top_type_by_segment[code] = None if pd.isna(vid) else str(vid).strip()

            coded.append({
                "source_index": si,
                "code": code,
                "name": name,
                "segments": segments,
                "section_type": top_type_by_segment.get(segments[0]),
                "outline": outline_by_si.get(si, 0),
            })
            last_node_code = code

        # Позиционный резолв предков: стек по глубине-кода (НЕ усечение кода).
        # Обеспечивает: дубли разводятся (берём ближайшего), forward-ref невозможен.
        depths = [len(c["segments"]) for c in coded]
        chains = resolve_ancestor_indices(depths)
        nodes: list[EstimateNode] = []
        for i, c in enumerate(coded):
            parts = [coded[j]["name"] for j in chains[i]]
            parts.append(c["name"])
            embedding_input = ". ".join(parts)  # байт-в-байт как template_parser
            nodes.append(EstimateNode(
                code=c["code"],
                name=c["name"],
                parent_code=".".join(c["segments"][:-1]) or None,  # код-based, не трогаем
                section_type=c["section_type"],
                embedding_input=embedding_input,
                source_index=c["source_index"],
                depth=len(c["segments"]),
            ))

        anomalies, overrides = detect_structural_anomalies(
            [(c["source_index"], c["code"], c["name"], c["outline"]) for c in coded]
        )
        # Если файл не содержит outline-группировки вовсе (все уровни == 0) —
        # outline-данных нет, overrides = 0 (не ошибка структуры, просто плоский файл).
        file_has_outline = any(c["outline"] > 0 for c in coded)
        if not file_has_outline:
            overrides = 0
        if outline_desync:  # рассинхрон → outline-детекция недостоверна; крошка/аномалии код-based — ок  # noqa: E501
            warnings.append("outline-детекция отключена: рассинхрон pandas↔openpyxl по строкам")
            overrides = 0

        return ParsedEstimate(
            nodes=nodes,
            positions=positions,
            warnings=warnings,
            anomalies=anomalies,
            outline_overrides=overrides,
        )

    @staticmethod
    def _read_outline(content: bytes) -> tuple[dict[int, int], dict[int, object]]:
        """Читает outline-уровни строк и коды из первого листа через openpyxl.

        НЕ read_only: нужен доступ к row_dimensions.outline_level.
        Возвращает (outline_by_si, code_by_si) где si — 0-based строка данных (как df.iterrows).
        """
        wb = load_workbook(io.BytesIO(content))  # НЕ read_only: нужен row_dimensions.outline_level
        ws = wb.worksheets[0]  # тот же лист, что pandas read_excel (sheet_name=0), НЕ wb.active
        header = [cell.value for cell in ws[1]]
        try:
            col = header.index(SECTION_NO_COLUMN) + 1
        except ValueError:  # pandas уже бы упал; перестраховка
            return {}, {}
        outline_by_si: dict[int, int] = {}
        code_by_si: dict[int, object] = {}
        for er in range(2, ws.max_row + 1):
            si = er - 2
            outline_by_si[si] = ws.row_dimensions[er].outline_level
            code_by_si[si] = ws.cell(row=er, column=col).value
        return outline_by_si, code_by_si
