"""Сценарий выгрузки сметы в .xlsx (SP3): оригинал из MinIO → заполнить «Статья СМР».

Пишем код только в строки-узлы по физ.строке source_index+2 (инвариант SP1: заголовок в
строке 1). Правило значения — см. спеку §5. Позиции-листья не трогаем (как в оригинале).
"""

from __future__ import annotations

from io import BytesIO

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.entities import StoredEstimateRow
from app.domain.errors import InvalidReviewActionError
from app.domain.ports import EstimateRepository, ObjectStorage

_HEADER = "статья смр"  # нормализованный заголовок-приёмник


class EstimateExportService:
    def __init__(self, estimates: EstimateRepository, storage: ObjectStorage) -> None:
        self._estimates = estimates
        self._storage = storage

    def export(
        self, estimate_id: int, requester_id: int, *, is_admin: bool, strict: bool = False
    ) -> bytes:
        key = self._estimates.get_object_key(estimate_id, requester_id, is_admin=is_admin)
        if key is None:
            raise LookupError("Смета не найдена")
        est = self._estimates.get(estimate_id, requester_id, is_admin=is_admin)
        if est is None:  # смета удалена между get_object_key и get (гонка) → 404, не 500
            raise LookupError("Смета не найдена")
        if strict:
            unreviewed = [
                r for r in est.rows
                if r.review_status == "unreviewed" and r.status in ("needs_review", "no_match")
            ]
            if unreviewed:
                raise InvalidReviewActionError(
                    f"Не просмотрено строк: {len(unreviewed)}"
                )
        raw = self._storage.get(key)  # сбой MinIO → StorageError долетит до роута → 503
        wb = openpyxl.load_workbook(BytesIO(raw))
        ws = wb.active
        col = self._find_or_create_column(ws)
        for row in est.rows:
            # пишем ВСЕГДА (в т.ч. "" для blank-случаев) — затирает старое значение
            # приёмника при повторном экспорте/шаблоне с примерами.
            ws.cell(row=row.source_index + 2, column=col, value=self._cell_value(row))
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    @staticmethod
    def _find_or_create_column(ws: Worksheet) -> int:
        for cell in ws[1]:
            if cell.value is not None and str(cell.value).strip().casefold() == _HEADER:
                return cell.column
        col = ws.max_column + 1
        ws.cell(row=1, column=col, value="Статья СМР")
        return col

    @staticmethod
    def _cell_value(row: StoredEstimateRow) -> str:
        if row.status == "excluded":
            return ""  # орг-заголовок исключён из матчинга → «Статья СМР» всегда пуста
        if row.review_status in ("confirmed", "overridden"):
            # для confirmed final_* = matched_* (заморожен в правке)
            return _format_article(row.final_code, row.final_name)
        if row.review_status == "unreviewed" and row.status in ("confident", "matched_fund"):
            # matched_fund авто-подтверждается только на фронте (review_status в БД unreviewed)
            return _format_article(row.matched_code, row.matched_name)
        return ""  # rejected ИЛИ unreviewed + needs_review/no_match/error/pending → пусто


def _format_article(code: str | None, name: str | None) -> str:
    """Ячейка «Статья СМР»: «(код) Название»; без имени — голый код, без кода — пусто."""
    if not code:
        return ""
    return f"({code}) {name}" if name else code
